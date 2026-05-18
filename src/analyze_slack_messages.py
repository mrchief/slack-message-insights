#!/usr/bin/env python3
"""Extract user-authored Slack conversations with context into Excel and JSONL."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_USER_ID = "U05BV05GYQH"


@dataclass(frozen=True)
class Message:
    channel_id: str
    channel_name: str
    ts: str
    thread_ts: str
    user: str
    text: str
    subtype: str
    datetime_local: datetime
    is_thread_root: bool
    attachments: str
    files: str
    reactions: str
    raw: dict[str, Any]

    @property
    def sort_key(self) -> tuple[str, float]:
        return (self.channel_id, slack_ts_to_float(self.ts))


def slack_ts_to_float(ts: str | None) -> float:
    if not ts:
        return 0.0
    try:
        return float(ts)
    except (TypeError, ValueError):
        return 0.0


def slack_ts_to_datetime(ts: str, tz: ZoneInfo) -> datetime:
    return datetime.fromtimestamp(slack_ts_to_float(ts), timezone.utc).astimezone(tz)


def parse_end_date(value: str | None, tz: ZoneInfo) -> datetime:
    if value:
        parsed = date.fromisoformat(value)
        return datetime.combine(parsed, time.max, tzinfo=tz)
    now = datetime.now(tz)
    return datetime.combine(now.date(), time.max, tzinfo=tz)


def load_target_channels(path: Path) -> list[str]:
    if not path.exists():
        return []
    ids = re.findall(r"\bC[A-Z0-9]{8,}\b", path.read_text(encoding="utf-8"))
    seen: set[str] = set()
    ordered: list[str] = []
    for channel_id in ids:
        if channel_id not in seen:
            seen.add(channel_id)
            ordered.append(channel_id)
    return ordered


def load_channel_inventory(path: Path) -> dict[str, dict[str, str]]:
    inventory: dict[str, dict[str, str]] = {}
    if not path.exists():
        return inventory
    pattern = re.compile(r"^(C[A-Z0-9]+)\s+(\S+)\s+(.+?)\s*$")
    for line in path.read_text(encoding="utf-8").splitlines():
        match = pattern.match(line)
        if not match:
            continue
        channel_id, archive_marker, label = match.groups()
        inventory[channel_id] = {
            "archived": "yes" if archive_marker == "arch" else "no",
            "label": label,
        }
    return inventory


def summarize_attachments(message: dict[str, Any]) -> str:
    parts: list[str] = []
    for attachment in message.get("attachments") or []:
        title = attachment.get("title") or attachment.get("fallback") or ""
        text = attachment.get("text") or ""
        link = attachment.get("title_link") or attachment.get("from_url") or attachment.get("original_url") or ""
        summary = " | ".join(part for part in [title, text, link] if part)
        if summary:
            parts.append(summary)
    return "\n".join(parts)


def summarize_files(message: dict[str, Any]) -> str:
    parts: list[str] = []
    for file_info in message.get("files") or []:
        title = file_info.get("title") or file_info.get("name") or file_info.get("id") or ""
        filetype = file_info.get("filetype") or file_info.get("pretty_type") or ""
        permalink = file_info.get("permalink") or ""
        summary = " | ".join(part for part in [title, filetype, permalink] if part)
        if summary:
            parts.append(summary)
    return "\n".join(parts)


def summarize_reactions(message: dict[str, Any]) -> str:
    parts = []
    for reaction in message.get("reactions") or []:
        name = reaction.get("name", "")
        count = reaction.get("count", 0)
        users = reaction.get("users") or []
        parts.append(f"{name}:{count} ({', '.join(users)})")
    return "\n".join(parts)


def make_message(raw: dict[str, Any], channel_id: str, channel_name: str, tz: ZoneInfo) -> Message | None:
    ts = str(raw.get("ts") or "")
    if not ts:
        return None
    thread_ts = str(raw.get("thread_ts") or "")
    return Message(
        channel_id=channel_id,
        channel_name=channel_name,
        ts=ts,
        thread_ts=thread_ts,
        user=str(raw.get("user") or raw.get("bot_id") or ""),
        text=str(raw.get("text") or ""),
        subtype=str(raw.get("subtype") or ""),
        datetime_local=slack_ts_to_datetime(ts, tz),
        is_thread_root=bool(raw.get("reply_count")) or bool(raw.get("slackdump_thread_replies")),
        attachments=summarize_attachments(raw),
        files=summarize_files(raw),
        reactions=summarize_reactions(raw),
        raw=raw,
    )


def load_dump(path: Path, tz: ZoneInfo) -> tuple[str, str, list[Message]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    channel_id = data.get("channel_id") or path.stem
    channel_name = data.get("name") or channel_id
    by_ts: dict[str, Message] = {}

    for raw in data.get("messages") or []:
        msg = make_message(raw, channel_id, channel_name, tz)
        if msg:
            by_ts[msg.ts] = msg
        for reply in raw.get("slackdump_thread_replies") or []:
            reply_msg = make_message(reply, channel_id, channel_name, tz)
            if reply_msg:
                by_ts[reply_msg.ts] = reply_msg

    messages = sorted(by_ts.values(), key=lambda message: slack_ts_to_float(message.ts))
    return channel_id, channel_name, messages


def message_in_window(message: Message, start: datetime, end: datetime) -> bool:
    return start <= message.datetime_local <= end


def top_level_anchor(message: Message) -> bool:
    return not message.thread_ts or message.thread_ts == message.ts


def root_ts(message: Message) -> str:
    return message.thread_ts or message.ts


def topic_hint(messages: list[Message]) -> str:
    for message in messages:
        text = compact_text(message.text) or compact_text(message.attachments)
        if text:
            return text[:180]
    return "(no text)"


def compact_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def transcript_line(message: Message, user_id: str) -> str:
    marker = "me" if message.user == user_id else message.user or "unknown"
    text = compact_text(message.text)
    extras = []
    if message.attachments:
        extras.append(f"attachments: {compact_text(message.attachments)}")
    if message.files:
        extras.append(f"files: {compact_text(message.files)}")
    suffix = f" [{' ; '.join(extras)}]" if extras else ""
    return f"[{message.datetime_local.isoformat(timespec='minutes')}] {marker}: {text}{suffix}"


def build_thread_groups(
    channel_id: str,
    messages: list[Message],
    user_id: str,
    start: datetime,
    end: datetime,
) -> tuple[list[dict[str, Any]], set[str]]:
    by_thread: dict[str, list[Message]] = defaultdict(list)
    for message in messages:
        if message.thread_ts:
            by_thread[root_ts(message)].append(message)
        elif message.is_thread_root:
            by_thread[message.ts].append(message)

    groups: list[dict[str, Any]] = []
    consumed_my_ts: set[str] = set()
    by_ts = {message.ts: message for message in messages}
    for thread_root, thread_messages in by_thread.items():
        if thread_root in by_ts and all(message.ts != thread_root for message in thread_messages):
            thread_messages.append(by_ts[thread_root])
        ordered = sorted({message.ts: message for message in thread_messages}.values(), key=lambda m: slack_ts_to_float(m.ts))
        my_messages = [m for m in ordered if m.user == user_id and message_in_window(m, start, end)]
        if not my_messages:
            continue
        consumed_my_ts.update(m.ts for m in my_messages)
        groups.append(make_group(channel_id, f"{channel_id}:thread:{thread_root}", "thread", ordered, my_messages, user_id))
    return groups, consumed_my_ts


def build_top_level_groups(
    channel_id: str,
    messages: list[Message],
    user_id: str,
    start: datetime,
    end: datetime,
    consumed_my_ts: set[str],
    context_before: int,
    context_after: int,
    merge_gap: timedelta,
) -> list[dict[str, Any]]:
    top_messages = [message for message in messages if top_level_anchor(message)]
    my_indexes = [
        index
        for index, message in enumerate(top_messages)
        if message.user == user_id and message.ts not in consumed_my_ts and message_in_window(message, start, end)
    ]
    if not my_indexes:
        return []

    windows: list[tuple[int, int]] = []
    for index in my_indexes:
        windows.append((max(0, index - context_before), min(len(top_messages) - 1, index + context_after)))

    merged: list[tuple[int, int]] = []
    for current_start, current_end in windows:
        if not merged:
            merged.append((current_start, current_end))
            continue
        previous_start, previous_end = merged[-1]
        previous_dt = top_messages[previous_end].datetime_local
        current_dt = top_messages[current_start].datetime_local
        if current_start <= previous_end + 1 or current_dt - previous_dt <= merge_gap:
            merged[-1] = (previous_start, max(previous_end, current_end))
        else:
            merged.append((current_start, current_end))

    groups: list[dict[str, Any]] = []
    for group_index, (start_index, end_index) in enumerate(merged, start=1):
        grouped_messages = top_messages[start_index : end_index + 1]
        my_messages = [m for m in grouped_messages if m.user == user_id and message_in_window(m, start, end)]
        group_id = f"{channel_id}:context:{group_index}:{grouped_messages[0].ts}"
        groups.append(make_group(channel_id, group_id, "context_window", grouped_messages, my_messages, user_id))
    return groups


def make_group(
    channel_id: str,
    group_id: str,
    group_type: str,
    messages: list[Message],
    my_messages: list[Message],
    user_id: str,
) -> dict[str, Any]:
    ordered = sorted({message.ts: message for message in messages}.values(), key=lambda m: slack_ts_to_float(m.ts))
    participants = sorted({message.user for message in ordered if message.user})
    transcript = "\n".join(transcript_line(message, user_id) for message in ordered)
    return {
        "group_id": group_id,
        "group_type": group_type,
        "channel_id": channel_id,
        "channel_name": ordered[0].channel_name if ordered else channel_id,
        "start": ordered[0].datetime_local if ordered else None,
        "end": ordered[-1].datetime_local if ordered else None,
        "message_count": len(ordered),
        "my_message_count": len(my_messages),
        "participants": participants,
        "topic_hint": topic_hint(ordered),
        "my_messages": my_messages,
        "messages": ordered,
        "transcript": transcript,
    }


def build_groups_by_channel(
    channel_id: str,
    messages: list[Message],
    user_id: str,
    start: datetime,
    end: datetime,
    context_before: int,
    context_after: int,
    merge_gap: timedelta,
) -> list[dict[str, Any]]:
    thread_groups, consumed_my_ts = build_thread_groups(channel_id, messages, user_id, start, end)
    context_groups = build_top_level_groups(
        channel_id,
        messages,
        user_id,
        start,
        end,
        consumed_my_ts,
        context_before,
        context_after,
        merge_gap,
    )
    return sorted(thread_groups + context_groups, key=lambda group: group["start"] or datetime.min.replace(tzinfo=timezone.utc))


def autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[column_letter].width = max(12, max_length + 2)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
    ws.freeze_panes = "A2"


def write_workbook(
    workbook_path: Path,
    groups: list[dict[str, Any]],
    summary_rows: list[tuple[str, Any]],
    warnings: list[str],
    user_id: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for key, value in summary_rows:
        ws.append([key, value])
    if warnings:
        ws.append(["Warnings", "\n".join(warnings)])
    autosize_sheet(ws)

    ws = wb.create_sheet("Conversations")
    ws.append(
        [
            "group_id",
            "group_type",
            "channel_id",
            "channel_name",
            "start",
            "end",
            "message_count",
            "my_message_count",
            "participants",
            "topic_hint",
            "my_messages",
            "transcript",
        ]
    )
    for group in groups:
        ws.append(
            [
                group["group_id"],
                group["group_type"],
                group["channel_id"],
                group["channel_name"],
                group["start"].isoformat(timespec="seconds") if group["start"] else "",
                group["end"].isoformat(timespec="seconds") if group["end"] else "",
                group["message_count"],
                group["my_message_count"],
                ", ".join(group["participants"]),
                group["topic_hint"],
                "\n".join(transcript_line(message, user_id) for message in group["my_messages"]),
                group["transcript"],
            ]
        )
    autosize_sheet(ws)

    ws = wb.create_sheet("Messages")
    ws.append(
        [
            "group_id",
            "group_type",
            "channel_id",
            "channel_name",
            "datetime",
            "ts",
            "thread_ts",
            "user",
            "is_me",
            "subtype",
            "text",
            "attachments",
            "files",
            "reactions",
        ]
    )
    for group in groups:
        for message in group["messages"]:
            ws.append(
                [
                    group["group_id"],
                    group["group_type"],
                    message.channel_id,
                    message.channel_name,
                    message.datetime_local.isoformat(timespec="seconds"),
                    message.ts,
                    message.thread_ts,
                    message.user,
                    "yes" if message.user == user_id else "no",
                    message.subtype,
                    message.text,
                    message.attachments,
                    message.files,
                    message.reactions,
                ]
            )
    autosize_sheet(ws)

    ws = wb.create_sheet("My Messages")
    ws.append(["group_id", "channel_id", "channel_name", "datetime", "ts", "thread_ts", "text", "attachments", "files"])
    for group in groups:
        for message in group["my_messages"]:
            ws.append(
                [
                    group["group_id"],
                    message.channel_id,
                    message.channel_name,
                    message.datetime_local.isoformat(timespec="seconds"),
                    message.ts,
                    message.thread_ts,
                    message.text,
                    message.attachments,
                    message.files,
                ]
            )
    autosize_sheet(ws)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def write_jsonl(path: Path, groups: list[dict[str, Any]], user_id: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for group in groups:
            payload = {
                "group_id": group["group_id"],
                "group_type": group["group_type"],
                "channel_id": group["channel_id"],
                "channel_name": group["channel_name"],
                "start": group["start"].isoformat(timespec="seconds") if group["start"] else "",
                "end": group["end"].isoformat(timespec="seconds") if group["end"] else "",
                "participants": group["participants"],
                "topic_hint": group["topic_hint"],
                "my_message_count": group["my_message_count"],
                "message_count": group["message_count"],
                "transcript": group["transcript"],
                "messages": [
                    {
                        "datetime": message.datetime_local.isoformat(timespec="seconds"),
                        "ts": message.ts,
                        "thread_ts": message.thread_ts,
                        "user": message.user,
                        "is_me": message.user == user_id,
                        "text": message.text,
                        "attachments": message.attachments,
                        "files": message.files,
                        "reactions": message.reactions,
                    }
                    for message in group["messages"]
                ],
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--user-id", default=DEFAULT_USER_ID)
    parser.add_argument("--channels-file", type=Path, default=ROOT / "info" / "channel list.txt")
    parser.add_argument("--channel-inventory", type=Path, default=ROOT / "info" / "channels-T01E0D9F1NW.txt")
    parser.add_argument("--dumps-dir", type=Path, default=ROOT / "dumps")
    parser.add_argument("--output-xlsx", type=Path, default=ROOT / "output" / "slack_message_analysis.xlsx")
    parser.add_argument("--output-jsonl", type=Path, default=ROOT / "output" / "conversation_contexts.jsonl")
    parser.add_argument("--timezone", default="America/New_York")
    parser.add_argument("--end-date", help="Inclusive end date in YYYY-MM-DD format. Defaults to today.")
    parser.add_argument("--months-days", type=int, default=90, help="Lookback window in days.")
    parser.add_argument("--context-before", type=int, default=4)
    parser.add_argument("--context-after", type=int, default=4)
    parser.add_argument("--merge-gap-minutes", type=int, default=30)
    parser.add_argument("--all-dumps", action="store_true", help="Analyze every dump file instead of only target channels.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    tz = ZoneInfo(args.timezone)
    end = parse_end_date(args.end_date, tz)
    start = end - timedelta(days=args.months_days)

    target_channels = load_target_channels(args.channels_file)
    inventory = load_channel_inventory(args.channel_inventory)
    dump_paths = {path.stem: path for path in sorted(args.dumps_dir.glob("*.json"))}
    selected_channel_ids = sorted(dump_paths) if args.all_dumps or not target_channels else target_channels

    warnings: list[str] = []
    missing_dumps = [channel_id for channel_id in selected_channel_ids if channel_id not in dump_paths]
    if missing_dumps:
        warnings.append(f"Missing dump files for {len(missing_dumps)} target channels: {', '.join(missing_dumps[:25])}")
        if len(missing_dumps) > 25:
            warnings.append(f"{len(missing_dumps) - 25} additional target channels are missing dumps.")

    all_groups: list[dict[str, Any]] = []
    loaded_channels = 0
    total_messages = 0
    empty_channels = 0
    my_authored_messages = 0

    for channel_id in selected_channel_ids:
        path = dump_paths.get(channel_id)
        if not path:
            continue
        loaded_channel_id, dump_channel_name, messages = load_dump(path, tz)
        loaded_channels += 1
        total_messages += len(messages)
        if not messages:
            empty_channels += 1
        display_name = inventory.get(loaded_channel_id, {}).get("label", dump_channel_name)
        normalized_messages = [
            Message(
                channel_id=message.channel_id,
                channel_name=display_name,
                ts=message.ts,
                thread_ts=message.thread_ts,
                user=message.user,
                text=message.text,
                subtype=message.subtype,
                datetime_local=message.datetime_local,
                is_thread_root=message.is_thread_root,
                attachments=message.attachments,
                files=message.files,
                reactions=message.reactions,
                raw=message.raw,
            )
            for message in messages
        ]
        my_authored_messages += sum(
            1 for message in normalized_messages if message.user == args.user_id and message_in_window(message, start, end)
        )
        all_groups.extend(
            build_groups_by_channel(
                loaded_channel_id,
                normalized_messages,
                args.user_id,
                start,
                end,
                args.context_before,
                args.context_after,
                timedelta(minutes=args.merge_gap_minutes),
            )
        )

    all_groups.sort(key=lambda group: group["start"] or datetime.min.replace(tzinfo=tz))
    extracted_message_count = sum(group["message_count"] for group in all_groups)
    extracted_my_message_count = sum(group["my_message_count"] for group in all_groups)

    summary_rows = [
        ("user_id", args.user_id),
        ("timezone", args.timezone),
        ("start", start.isoformat(timespec="seconds")),
        ("end", end.isoformat(timespec="seconds")),
        ("target_channels", len(selected_channel_ids)),
        ("loaded_channels", loaded_channels),
        ("empty_channels", empty_channels),
        ("available_dump_files", len(dump_paths)),
        ("total_loaded_messages", total_messages),
        ("authored_messages_in_window", my_authored_messages),
        ("conversation_groups", len(all_groups)),
        ("extracted_messages_with_context", extracted_message_count),
        ("extracted_authored_messages", extracted_my_message_count),
        ("context_before", args.context_before),
        ("context_after", args.context_after),
        ("merge_gap_minutes", args.merge_gap_minutes),
        ("channels_file", str(args.channels_file)),
        ("channel_inventory", str(args.channel_inventory)),
        ("dumps_dir", str(args.dumps_dir)),
    ]

    write_workbook(args.output_xlsx, all_groups, summary_rows, warnings, args.user_id)
    write_jsonl(args.output_jsonl, all_groups, args.user_id)

    print(f"Wrote {args.output_xlsx}")
    print(f"Wrote {args.output_jsonl}")
    print(f"Conversation groups: {len(all_groups)}")
    print(f"Extracted messages with context: {extracted_message_count}")
    print(f"Authored messages extracted: {extracted_my_message_count}")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
